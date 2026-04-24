"""Web of Science publication data parser.

Parses Victor's WoS export Excel file with 3 tabs:
  Tab 1: Publications with keywords and metadata (6,019 pubs)
  Tab 2: Deduplicated Keywords Plus vocabulary (7,488 terms)
  Tab 3: NETL organizational/technology taxonomy (3,877 pubs)
"""

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)

WOS_EXCEL_PATH = Path("data/WoS/WOS_research_keywords_2.xlsx")


def _split_comma_field(value: str | None) -> list[str]:
    if not value or not str(value).strip():
        return []
    return [k.strip() for k in str(value).split(",") if k.strip()]


def _clean_html_entities(text: str | None) -> str | None:
    if not text:
        return text
    text = re.sub(r"</?sub>", "", str(text))
    text = re.sub(r"</?sup>", "", text)
    text = re.sub(r"</?[a-zA-Z]+[^>]*>", "", text)
    return text.strip()


def parse_tab1_publications(excel_path: Path = WOS_EXCEL_PATH) -> list[dict]:
    """Parse Tab 1: publications with keywords and WoS metadata."""
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb["keywords with Pub IDS"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = rows[0]
    data = rows[1:]
    records = []

    for row in data:
        acc = row[0]
        if not acc or not str(acc).strip():
            continue

        records.append({
            "accession_number": str(acc).strip(),
            "keywords_author": _split_comma_field(row[1]),
            "keywords_plus": _split_comma_field(row[2]),
            "subject_sub_heading_1": str(row[3]).strip() if row[3] else None,
            "subject_sub_heading_2": str(row[4]).strip() if row[4] else None,
            "subject_cat_traditional_1": str(row[5]).strip() if row[5] else None,
            "subject_cat_traditional_2": str(row[6]).strip() if row[6] else None,
            "subject_cat_extended": _split_comma_field(row[7]),
            "category_heading_1": str(row[8]).strip() if row[8] else None,
            "category_heading_2": str(row[9]).strip() if row[9] else None,
            "abstract": _clean_html_entities(row[10]),
            "source_title": str(row[11]).strip() if row[11] else None,
            "title": _clean_html_entities(row[12]),
            "doc_type_1": str(row[13]).strip() if row[13] else None,
            "doc_type_2": str(row[14]).strip() if row[14] else None,
            "grant_agencies": _split_comma_field(row[15]),
            "data_acquired": row[16],
        })

    logger.info(f"Tab 1: parsed {len(records)} publications")
    return records


def parse_tab2_keywords_plus(excel_path: Path = WOS_EXCEL_PATH) -> list[dict]:
    """Parse Tab 2: deduplicated Keywords Plus vocabulary."""
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb["Keywords 2"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    for row in rows:
        val = row[0]
        if not val or not str(val).strip():
            continue
        kw = str(val).strip()
        if kw == "Keywords Plus2":
            continue
        cleaned = re.sub(r"[()]+$", "", kw).strip()
        records.append({
            "keyword": kw,
            "normalized": cleaned.lower(),
        })

    logger.info(f"Tab 2: parsed {len(records)} Keywords Plus terms")
    return records


def parse_tab3_netl_tech(excel_path: Path = WOS_EXCEL_PATH) -> list[dict]:
    """Parse Tab 3: NETL organizational/technology taxonomy."""
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb["Keywords 3"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = rows[0]
    data = rows[1:]
    records = []

    for row in data:
        title = row[0]
        if not title or not str(title).strip():
            continue

        def clean_na(val):
            if val is None:
                return None
            s = str(val).strip()
            if s in ("#N/A", "0", ""):
                return None
            return s

        records.append({
            "article_title": _clean_html_entities(title),
            "technology_area": clean_na(row[1]),
            "program_area": clean_na(row[2]),
            "sub_program_area": clean_na(row[3]),
            "technology_area_alt": clean_na(row[4]),
            "consolidated_tech_area": clean_na(row[5]),
            "consolidated_tech_filter": clean_na(row[6]),
            "turbines_sub_tech": clean_na(row[7]),
        })

    logger.info(f"Tab 3: parsed {len(records)} NETL tech records")
    return records


def ingest_all(conn, excel_path: Path = WOS_EXCEL_PATH) -> dict:
    """Ingest all 3 WoS tabs into DuckDB staging tables."""
    from ..schema import RAW_WOS_PUBLICATIONS, RAW_WOS_KEYWORDS_PLUS, RAW_WOS_NETL_TECH
    from ..schema import _strip_sql_comments

    for ddl in [RAW_WOS_PUBLICATIONS, RAW_WOS_KEYWORDS_PLUS, RAW_WOS_NETL_TECH]:
        cleaned = _strip_sql_comments(ddl)
        for stmt in cleaned.split(";"):
            stmt = stmt.strip()
            if stmt:
                conn.execute(stmt)

    stats = {}

    tab1 = parse_tab1_publications(excel_path)
    if tab1:
        conn.execute("DELETE FROM raw_wos_publications")
        conn.executemany(
            """INSERT INTO raw_wos_publications
               (accession_number, keywords_author, keywords_plus,
                subject_sub_heading_1, subject_sub_heading_2,
                subject_cat_traditional_1, subject_cat_traditional_2,
                subject_cat_extended, category_heading_1, category_heading_2,
                abstract, source_title, title, doc_type_1, doc_type_2,
                grant_agencies, data_acquired)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [(r["accession_number"], r["keywords_author"], r["keywords_plus"],
              r["subject_sub_heading_1"], r["subject_sub_heading_2"],
              r["subject_cat_traditional_1"], r["subject_cat_traditional_2"],
              r["subject_cat_extended"], r["category_heading_1"], r["category_heading_2"],
              r["abstract"], r["source_title"], r["title"],
              r["doc_type_1"], r["doc_type_2"],
              r["grant_agencies"], r["data_acquired"])
             for r in tab1]
        )
        stats["publications"] = len(tab1)

    tab2 = parse_tab2_keywords_plus(excel_path)
    if tab2:
        conn.execute("DELETE FROM raw_wos_keywords_plus_vocab")
        conn.executemany(
            "INSERT INTO raw_wos_keywords_plus_vocab (keyword, normalized) VALUES (?, ?)",
            [(r["keyword"], r["normalized"]) for r in tab2]
        )
        stats["keywords_plus_vocab"] = len(tab2)

    tab3 = parse_tab3_netl_tech(excel_path)
    if tab3:
        conn.execute("DELETE FROM raw_wos_netl_tech")
        conn.executemany(
            """INSERT INTO raw_wos_netl_tech
               (article_title, technology_area, program_area, sub_program_area,
                technology_area_alt, consolidated_tech_area, consolidated_tech_filter,
                turbines_sub_tech)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            [(r["article_title"], r["technology_area"], r["program_area"],
              r["sub_program_area"], r["technology_area_alt"],
              r["consolidated_tech_area"], r["consolidated_tech_filter"],
              r["turbines_sub_tech"])
             for r in tab3]
        )
        stats["netl_tech"] = len(tab3)

    return stats
