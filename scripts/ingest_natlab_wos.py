"""Ingest WoS national labs publication data and generate senses + relationships.

Handles the full pipeline:
  1. Parse Excel → raw_wos_natlab_publications table
  2. Generate keyword senses (WoS_natlab_pub source)
  3. Generate category/journal senses
  4. Cross-link to existing pillar senses
  5. Re-run relationship strategies on new senses
"""

import duckdb
import logging
import re
import sys
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent.parent))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_PATH = Path("data/WoS/WoS_other_NatLAbs_keyword_Pubs.xlsx")

# Column mapping: Excel col index → field name
COLUMNS = [
    "keywords_author",           # 0
    "keywords_plus",             # 1
    "accession_number",          # 2
    "category_heading_1",        # 3
    "doc_type_1",                # 4
    "subject_cat_traditional_1", # 5
    "subject_sub_heading_1",     # 6
    "category_heading_2",        # 7
    "doc_type_2",                # 8
    "subject_cat_traditional_2", # 9
    "subject_sub_heading_2",     # 10
    "doc_type_3",                # 11
    "abstract",                  # 12
    "category",                  # 13
    "publisher",                 # 14
    "source_title",              # 15
    "sub_category",              # 16
    "subject_category",          # 17
    "subject_cat_extended",      # 18
    "table_names",               # 19
    "title",                     # 20
]


def _split_semicolons(value: str | None) -> list[str]:
    """Split semicolon-delimited keywords (WoS format)."""
    if not value or not str(value).strip():
        return []
    return [k.strip() for k in str(value).split(";") if k.strip()]


def _clean_html(text: str | None) -> str | None:
    if not text:
        return text
    text = re.sub(r"</?[a-zA-Z]+[^>]*>", "", str(text))
    return text.strip() or None


def step1_parse_and_ingest(conn) -> int:
    """Parse Excel and insert into raw_wos_natlab_publications."""
    import openpyxl

    from src.schema import RAW_WOS_NATLAB, _strip_sql_comments
    cleaned = _strip_sql_comments(RAW_WOS_NATLAB)
    for stmt in cleaned.split(";"):
        stmt = stmt.strip()
        if stmt:
            conn.execute(stmt)

    logger.info("Reading Excel file (255K rows, streaming)...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Check for overlap with existing NETL WoS
    existing_uts = set()
    try:
        for r in conn.execute("SELECT accession_number FROM raw_wos_publications").fetchall():
            existing_uts.add(r[0])
        logger.info(f"Existing NETL WoS papers: {len(existing_uts)}")
    except Exception:
        pass

    conn.execute("DELETE FROM raw_wos_natlab_publications")

    conn.execute("""CREATE TEMP TABLE _tmp_natlab (
        accession_number VARCHAR,
        keywords_author VARCHAR,
        keywords_plus VARCHAR,
        category_heading_1 VARCHAR,
        doc_type_1 VARCHAR,
        subject_cat_traditional_1 VARCHAR,
        subject_sub_heading_1 VARCHAR,
        category_heading_2 VARCHAR,
        doc_type_2 VARCHAR,
        subject_cat_traditional_2 VARCHAR,
        subject_sub_heading_2 VARCHAR,
        doc_type_3 VARCHAR,
        abstract TEXT,
        category VARCHAR,
        publisher VARCHAR,
        source_title VARCHAR,
        sub_category VARCHAR,
        subject_category VARCHAR,
        subject_cat_extended VARCHAR,
        table_names VARCHAR,
        title VARCHAR
    )""")

    batch = []
    seen_uts = set()
    skipped_overlap = 0
    total = 0
    batch_size = 10000

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        ut = row[2]
        if not ut or not str(ut).strip():
            continue
        ut = str(ut).strip()

        if ut in existing_uts:
            skipped_overlap += 1
            continue
        if ut in seen_uts:
            continue
        seen_uts.add(ut)

        batch.append((
            ut,
            str(row[0]).strip() if row[0] else None,  # keywords raw string
            str(row[1]).strip() if row[1] else None,  # kw plus raw string
            str(row[3]).strip() if row[3] else None,
            str(row[4]).strip() if row[4] else None,
            str(row[5]).strip() if row[5] else None,
            str(row[6]).strip() if row[6] else None,
            str(row[7]).strip() if row[7] else None,
            str(row[8]).strip() if row[8] else None,
            str(row[9]).strip() if row[9] else None,
            str(row[10]).strip() if row[10] else None,
            str(row[11]).strip() if row[11] else None,
            _clean_html(row[12]),
            str(row[13]).strip() if row[13] else None,
            str(row[14]).strip() if row[14] else None,
            str(row[15]).strip() if row[15] else None,
            str(row[16]).strip() if row[16] else None,
            str(row[17]).strip() if row[17] else None,
            str(row[18]).strip() if row[18] else None,
            str(row[19]).strip() if row[19] else None,
            _clean_html(row[20]),
        ))
        total += 1

        if len(batch) >= batch_size:
            conn.executemany(
                "INSERT INTO _tmp_natlab VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                batch,
            )
            batch = []
            if total % 50000 == 0:
                logger.info(f"  ...{total:,} rows parsed")

    if batch:
        conn.executemany(
            "INSERT INTO _tmp_natlab VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            batch,
        )

    wb.close()
    logger.info(f"Parsed {total:,} unique papers ({skipped_overlap} overlap with NETL WoS skipped)")

    # Now insert with proper array splitting via SQL
    conn.execute("""
        INSERT INTO raw_wos_natlab_publications
            (accession_number, keywords_author, keywords_plus,
             category_heading_1, doc_type_1, subject_cat_traditional_1,
             subject_sub_heading_1, category_heading_2, doc_type_2,
             subject_cat_traditional_2, subject_sub_heading_2, doc_type_3,
             abstract, category, publisher, source_title, sub_category,
             subject_category, subject_cat_extended, table_names, title)
        SELECT
            accession_number,
            CASE WHEN keywords_author IS NOT NULL
                 THEN string_split(keywords_author, '; ')
                 ELSE NULL END,
            CASE WHEN keywords_plus IS NOT NULL
                 THEN string_split(keywords_plus, '; ')
                 ELSE NULL END,
            category_heading_1, doc_type_1, subject_cat_traditional_1,
            subject_sub_heading_1, category_heading_2, doc_type_2,
            subject_cat_traditional_2, subject_sub_heading_2, doc_type_3,
            abstract, category, publisher, source_title, sub_category,
            subject_category, subject_cat_extended, table_names, title
        FROM _tmp_natlab
    """)
    conn.execute("DROP TABLE _tmp_natlab")

    count = conn.execute("SELECT COUNT(*) FROM raw_wos_natlab_publications").fetchone()[0]
    logger.info(f"Inserted {count:,} papers into raw_wos_natlab_publications")

    # Quick stats
    kw_count = conn.execute(
        "SELECT COUNT(*) FROM raw_wos_natlab_publications WHERE keywords_author IS NOT NULL AND array_length(keywords_author) > 0"
    ).fetchone()[0]
    kwp_count = conn.execute(
        "SELECT COUNT(*) FROM raw_wos_natlab_publications WHERE keywords_plus IS NOT NULL AND array_length(keywords_plus) > 0"
    ).fetchone()[0]
    abs_count = conn.execute(
        "SELECT COUNT(*) FROM raw_wos_natlab_publications WHERE abstract IS NOT NULL"
    ).fetchone()[0]
    logger.info(f"  With author keywords: {kw_count:,}")
    logger.info(f"  With Keywords Plus: {kwp_count:,}")
    logger.info(f"  With abstracts: {abs_count:,}")

    return count


def step2_generate_senses(conn) -> dict:
    """Generate keyword senses from natlab WoS publication keywords.

    Creates senses with origin_source = 'WoS_natlab_pub' and cross-links
    to existing pillar senses where labels match.
    """
    from src.ontology import _classify_wos_keyword, WOS_CAT_TO_DISCIPLINE

    conn.execute("DELETE FROM keyword_senses WHERE origin_source = 'WoS_natlab_pub'")

    # Load existing labels and sense IDs for dedup and cross-linking
    existing_labels = set()
    existing_sense_ids = set()
    for (sid,) in conn.execute("SELECT sense_id FROM keyword_senses").fetchall():
        existing_sense_ids.add(sid)
    for (label,) in conn.execute("SELECT DISTINCT LOWER(keyword_label) FROM keyword_senses").fetchall():
        existing_labels.add(label)

    logger.info(f"Existing senses: {len(existing_sense_ids):,}, unique labels: {len(existing_labels):,}")

    # Collect all keywords from natlab pubs
    pubs = conn.execute("""
        SELECT accession_number, keywords_author, keywords_plus,
               subject_cat_traditional_1, source_title, title,
               category, sub_category, subject_category,
               subject_sub_heading_1, subject_sub_heading_2,
               doc_type_1
        FROM raw_wos_natlab_publications
    """).fetchall()

    keyword_counts = Counter()
    keyword_cats = {}  # keyword → most common category

    for pub in pubs:
        acc, kw_auth, kw_plus, cat1, source, title, category, subcat, subj_cat, sub1, sub2, dt1 = pub
        all_kw = set()
        for kw_list in [kw_auth, kw_plus]:
            if kw_list:
                for kw in kw_list:
                    cleaned = kw.strip().lower()
                    if len(cleaned) > 2:
                        all_kw.add(cleaned)

        # Category fields are also keyword-capable
        for field in [cat1, category, subcat, subj_cat, sub1, sub2]:
            if field and field.strip() and len(field.strip()) > 2:
                all_kw.add(field.strip().lower())

        for kw in all_kw:
            keyword_counts[kw] += 1
            if kw not in keyword_cats and cat1:
                keyword_cats[kw] = cat1

    logger.info(f"Unique keywords from natlab pubs: {len(keyword_counts):,}")

    # Split into matching vs new
    matching = {kw for kw in keyword_counts if kw in existing_labels}
    new_kw = {kw for kw in keyword_counts if kw not in existing_labels}
    logger.info(f"  Match existing senses: {len(matching):,}")
    logger.info(f"  New keywords: {len(new_kw):,}")

    # Generate senses for new keywords only
    senses = []
    new_rels = []
    new_sense_ids = set()

    for kw in sorted(new_kw):
        normalized = kw.replace(" ", "_")[:60]
        sense_id = f"wosnl:{normalized}@WoS_natlab_pub#0"
        if sense_id in existing_sense_ids or sense_id in new_sense_ids:
            continue
        new_sense_ids.add(sense_id)

        cat = keyword_cats.get(kw)
        disc = _classify_wos_keyword(kw, cat)
        freq = keyword_counts[kw]

        from src.ontology import _get_resolution_tier
        tier_map = dict(conn.execute("SELECT discipline_id, tier FROM disciplines").fetchall())

        tags = []
        if freq >= 10:
            tags.append("high_frequency")
        tier = _get_resolution_tier(disc, tier_map)

        senses.append((
            sense_id, f"wosnl:{normalized}", "WoS_natlab_pub", kw,
            "WoS_natlab_pub", None, None,
            disc, [], tier,
            None, None, None,
            tags, min(0.95, 0.5 + freq / 100), "wos_natlab_extraction",
        ))

    # Cross-link matching keywords to existing senses
    if matching:
        from src.ontology import _get_resolution_tier
        tier_map = dict(conn.execute("SELECT discipline_id, tier FROM disciplines").fetchall())

        existing_by_label = {}
        for sid, label in conn.execute(
            "SELECT sense_id, LOWER(keyword_label) FROM keyword_senses"
        ).fetchall():
            if label not in existing_by_label:
                existing_by_label[label] = []
            existing_by_label[label].append(sid)

        for kw in matching:
            targets = existing_by_label.get(kw, [])
            if not targets:
                continue
            normalized = kw.replace(" ", "_")[:60]
            sense_id = f"wosnl:{normalized}@WoS_natlab_pub#0"
            if sense_id in existing_sense_ids or sense_id in new_sense_ids:
                for target_sid in targets[:3]:
                    new_rels.append((sense_id, target_sid, "related_to", "toward", 0.8, "wos_natlab_label_match", []))
                continue
            new_sense_ids.add(sense_id)

            cat = keyword_cats.get(kw)
            disc = _classify_wos_keyword(kw, cat)
            freq = keyword_counts[kw]
            tier = _get_resolution_tier(disc, tier_map)

            tags = []
            if freq >= 10:
                tags.append("high_frequency")
            if kw in existing_labels:
                tags.append("pillar_overlap")

            senses.append((
                sense_id, f"wosnl:{normalized}", "WoS_natlab_pub", kw,
                "WoS_natlab_pub", None, None,
                disc, [], tier,
                None, None, None,
                tags, min(0.95, 0.5 + freq / 100), "wos_natlab_extraction",
            ))

            for target_sid in targets[:3]:
                new_rels.append((sense_id, target_sid, "related_to", "toward", 0.8, "wos_natlab_label_match", []))

    logger.info(f"New senses to insert: {len(senses):,}")
    logger.info(f"Cross-link relationships: {len(new_rels):,}")

    # Insert senses via temp table (avoid executemany VARCHAR[] pathology)
    if senses:
        # Avoid VARCHAR[] in executemany (DuckDB pathological perf).
        # Insert scalar columns only, add arrays via SQL UPDATE.
        conn.execute("""CREATE TEMP TABLE _tmp_nl_senses (
            sense_id VARCHAR, keyword_id VARCHAR, keyword_source VARCHAR,
            keyword_label VARCHAR, origin_source VARCHAR, origin_path VARCHAR,
            origin_level INTEGER, discipline_primary VARCHAR,
            resolution_tier INTEGER,
            definition_in_context VARCHAR, scope_note VARCHAR,
            disambiguation VARCHAR,
            confidence FLOAT, provenance VARCHAR,
            tags_csv VARCHAR
        )""")
        batch = 5000
        for i in range(0, len(senses), batch):
            chunk = senses[i:i + batch]
            conn.executemany(
                "INSERT INTO _tmp_nl_senses VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [(s[0], s[1], s[2], s[3], s[4], s[5], s[6], s[7],
                  s[9], s[10], s[11], s[12],
                  s[14], s[15],
                  ",".join(s[13]) if s[13] else None)
                 for s in chunk],
            )
        conn.execute("""
            INSERT INTO keyword_senses
                (sense_id, keyword_id, keyword_source, keyword_label, origin_source,
                 origin_path, origin_level, discipline_primary,
                 disciplines_secondary, resolution_tier, definition_in_context,
                 scope_note, disambiguation, relevance_tags, confidence, provenance)
            SELECT sense_id, keyword_id, keyword_source, keyword_label, origin_source,
                   origin_path, origin_level, discipline_primary,
                   ARRAY[]::VARCHAR[], resolution_tier, definition_in_context,
                   scope_note, disambiguation,
                   CASE WHEN tags_csv IS NOT NULL THEN string_split(tags_csv, ',')
                        ELSE ARRAY[]::VARCHAR[] END,
                   confidence, provenance
            FROM _tmp_nl_senses
        """)
        conn.execute("DROP TABLE _tmp_nl_senses")

    # Insert relationships
    if new_rels:
        conn.execute("""CREATE TEMP TABLE _tmp_nl_rels (
            source VARCHAR, target VARCHAR, rel_type VARCHAR,
            direction VARCHAR, confidence DOUBLE, provenance VARCHAR, lens VARCHAR[]
        )""")
        # Dedup
        seen = set()
        deduped = []
        for r in new_rels:
            key = (r[0], r[1], r[2])
            if key not in seen:
                deduped.append(r)
                seen.add(key)
        for i in range(0, len(deduped), 5000):
            conn.executemany(
                "INSERT INTO _tmp_nl_rels VALUES (?,?,?,?,?,?,?)",
                deduped[i:i + 5000],
            )
        conn.execute("""
            INSERT INTO sense_relationships
                (source_sense_id, target_sense_id, relationship_type,
                 direction, confidence, provenance, lens_contexts)
            SELECT source, target, rel_type, direction, confidence, provenance, lens
            FROM _tmp_nl_rels
            WHERE (source, target, rel_type) NOT IN (
                SELECT source_sense_id, target_sense_id, relationship_type
                FROM sense_relationships
            )
        """)
        conn.execute("DROP TABLE _tmp_nl_rels")

    total_senses = conn.execute("SELECT COUNT(*) FROM keyword_senses").fetchone()[0]
    natlab_senses = conn.execute(
        "SELECT COUNT(*) FROM keyword_senses WHERE origin_source = 'WoS_natlab_pub'"
    ).fetchone()[0]
    total_rels = conn.execute("SELECT COUNT(*) FROM sense_relationships").fetchone()[0]

    logger.info(f"Total senses now: {total_senses:,} (natlab: {natlab_senses:,})")
    logger.info(f"Total relationships now: {total_rels:,}")

    return {
        "new_senses": len(senses),
        "cross_links": len(deduped) if new_rels else 0,
        "total_senses": total_senses,
        "total_rels": total_rels,
    }


def step3_category_senses(conn) -> int:
    """Generate senses from natlab WoS category fields (unique values only)."""
    from src.ontology import _classify_wos_keyword

    existing_labels = set()
    existing_sense_ids = set()
    for (label,) in conn.execute("SELECT DISTINCT LOWER(keyword_label) FROM keyword_senses").fetchall():
        existing_labels.add(label)
    for (sid,) in conn.execute("SELECT sense_id FROM keyword_senses").fetchall():
        existing_sense_ids.add(sid)

    # Collect unique category values across all category-type columns
    categories = set()
    for col in ["category_heading_1", "category_heading_2", "subject_cat_traditional_1",
                "subject_cat_traditional_2", "subject_sub_heading_1", "subject_sub_heading_2",
                "category", "sub_category", "subject_category"]:
        for (val,) in conn.execute(f"""
            SELECT DISTINCT {col} FROM raw_wos_natlab_publications
            WHERE {col} IS NOT NULL AND LENGTH(TRIM({col})) > 2
        """).fetchall():
            categories.add(val.strip())

    # Unique publishers
    publishers = set()
    for (val,) in conn.execute("""
        SELECT DISTINCT publisher FROM raw_wos_natlab_publications
        WHERE publisher IS NOT NULL AND LENGTH(TRIM(publisher)) > 2
    """).fetchall():
        publishers.add(val.strip())

    # Unique doc types
    doc_types = set()
    for col in ["doc_type_1", "doc_type_2", "doc_type_3"]:
        for (val,) in conn.execute(f"""
            SELECT DISTINCT {col} FROM raw_wos_natlab_publications
            WHERE {col} IS NOT NULL AND LENGTH(TRIM({col})) > 2
        """).fetchall():
            doc_types.add(val.strip())

    logger.info(f"Unique categories: {len(categories)}, publishers: {len(publishers)}, doc types: {len(doc_types)}")

    senses = []
    for cat in sorted(categories):
        cat_lower = cat.lower()
        if cat_lower in existing_labels:
            continue
        normalized = cat_lower.replace(" ", "_").replace(",", "")[:60]
        sense_id = f"wosnlcat:{normalized}@WoS_natlab_cat#0"
        if sense_id in existing_sense_ids:
            continue
        disc = _classify_wos_keyword(cat_lower, cat)
        from src.ontology import _get_resolution_tier
        tier_map = dict(conn.execute("SELECT discipline_id, tier FROM disciplines").fetchall())
        tier = _get_resolution_tier(disc, tier_map)
        senses.append((
            sense_id, f"wosnlcat:{normalized}", "WoS_natlab_cat", cat,
            "WoS_natlab_cat", None, None, disc, [], tier,
            None, None, None, [], 0.6, "wos_natlab_category",
        ))

    for pub in sorted(publishers):
        pub_lower = pub.lower()
        if pub_lower in existing_labels:
            continue
        normalized = pub_lower.replace(" ", "_")[:60]
        sense_id = f"wosnlpub:{normalized}@WoS_natlab_publisher#0"
        if sense_id in existing_sense_ids:
            continue
        senses.append((
            sense_id, f"wosnlpub:{normalized}", "WoS_natlab_publisher", pub,
            "WoS_natlab_publisher", None, None, "policy_economics", [], 3,
            None, None, None, [], 0.5, "wos_natlab_publisher",
        ))

    for dt in sorted(doc_types):
        dt_lower = dt.lower()
        if dt_lower in existing_labels:
            continue
        normalized = dt_lower.replace(" ", "_")[:60]
        sense_id = f"wosnldt:{normalized}@WoS_natlab_doctype#0"
        if sense_id in existing_sense_ids:
            continue
        senses.append((
            sense_id, f"wosnldt:{normalized}", "WoS_natlab_doctype", dt,
            "WoS_natlab_doctype", None, None, "policy_economics", [], 4,
            None, None, None, [], 0.4, "wos_natlab_doctype",
        ))

    logger.info(f"New category/publisher/doctype senses: {len(senses)}")

    if senses:
        conn.execute("""CREATE TEMP TABLE _tmp_nl_meta (
            sense_id VARCHAR, keyword_id VARCHAR, keyword_source VARCHAR,
            keyword_label VARCHAR, origin_source VARCHAR, origin_path VARCHAR,
            origin_level INTEGER, discipline_primary VARCHAR,
            resolution_tier INTEGER,
            definition_in_context VARCHAR, scope_note VARCHAR,
            disambiguation VARCHAR,
            confidence FLOAT, provenance VARCHAR,
            tags_csv VARCHAR
        )""")
        for i in range(0, len(senses), 5000):
            chunk = senses[i:i + 5000]
            conn.executemany(
                "INSERT INTO _tmp_nl_meta VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [(s[0], s[1], s[2], s[3], s[4], s[5], s[6], s[7],
                  s[9], s[10], s[11], s[12],
                  s[14], s[15],
                  ",".join(s[13]) if s[13] else None)
                 for s in chunk],
            )
        conn.execute("""
            INSERT INTO keyword_senses
                (sense_id, keyword_id, keyword_source, keyword_label, origin_source,
                 origin_path, origin_level, discipline_primary,
                 disciplines_secondary, resolution_tier, definition_in_context,
                 scope_note, disambiguation, relevance_tags, confidence, provenance)
            SELECT sense_id, keyword_id, keyword_source, keyword_label, origin_source,
                   origin_path, origin_level, discipline_primary,
                   ARRAY[]::VARCHAR[], resolution_tier, definition_in_context,
                   scope_note, disambiguation,
                   CASE WHEN tags_csv IS NOT NULL THEN string_split(tags_csv, ',')
                        ELSE ARRAY[]::VARCHAR[] END,
                   confidence, provenance
            FROM _tmp_nl_meta
        """)
        conn.execute("DROP TABLE _tmp_nl_meta")

    return len(senses)


def main():
    conn = duckdb.connect("data/lake/keywords.duckdb")

    print("=" * 60)
    print("STEP 1: Parse & ingest natlab WoS publications")
    print("=" * 60)
    pub_count = step1_parse_and_ingest(conn)

    print("\n" + "=" * 60)
    print("STEP 2: Generate keyword senses + cross-links")
    print("=" * 60)
    sense_stats = step2_generate_senses(conn)

    print("\n" + "=" * 60)
    print("STEP 3: Generate category/publisher/doctype senses")
    print("=" * 60)
    meta_count = step3_category_senses(conn)

    # Final summary
    total_senses = conn.execute("SELECT COUNT(*) FROM keyword_senses").fetchone()[0]
    total_rels = conn.execute("SELECT COUNT(*) FROM sense_relationships").fetchone()[0]
    natlab_senses = conn.execute(
        "SELECT COUNT(*) FROM keyword_senses WHERE origin_source LIKE 'WoS_natlab%'"
    ).fetchone()[0]

    print("\n" + "=" * 60)
    print("NATLAB WoS INGESTION COMPLETE")
    print("=" * 60)
    print(f"Publications ingested: {pub_count:,}")
    print(f"New keyword senses: {sense_stats['new_senses']:,}")
    print(f"Category/pub/doctype senses: {meta_count:,}")
    print(f"Cross-link relationships: {sense_stats['cross_links']:,}")
    print(f"Total natlab senses: {natlab_senses:,}")
    print(f"Total senses (all sources): {total_senses:,}")
    print(f"Total relationships: {total_rels:,}")
    print(f"Rels/sense: {total_rels / total_senses:.3f}")

    print(f"\nSenses by origin_source:")
    for r in conn.execute("""
        SELECT origin_source, COUNT(*) FROM keyword_senses
        GROUP BY 1 ORDER BY 2 DESC
    """).fetchall():
        print(f"  {r[0]}: {r[1]:,}")

    conn.close()


if __name__ == "__main__":
    main()
